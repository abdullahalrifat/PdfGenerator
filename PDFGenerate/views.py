import xlrd
from django.http import HttpResponse, HttpResponseRedirect
from django.http import  Http404
from django.template import loader
from django.shortcuts import render
from django.views import  generic
from django.views.generic.edit import CreateView, UpdateView, DeleteView

from Test.settings import BASE_DIR
from Test.settings import STATIC_DIRS
from Test.settings import PDF_DIRS,MEDIA_URL,MEDIA_ROOT
from .models import pdf
from django.template.loader import get_template
from django.template import Context
from django.conf import settings


from .models import Comment
from .models import pdf
from .models import verify
from .models import excel
from .models import data

from .forms import AddCommentForm
from .forms import AddIdForm
from .forms import AddPasswordForm
from .forms import AddExcelForm


from openpyxl import load_workbook


import pdfkit

'''


def index(request):

    users=pdf.objects.all()
    context={
        'allUsers':users,
    }
        #pdfkit.from_file('PDFGenerate/templates/PDFGenerate/pdf_form.html', 'out.pdf')

    return render(request,'PDFGenerate/pdf_form.html',context)

def pdfgen(request):

    # Create a  URL of our project and go to the template route
    projectUrl = request.get_host() + '/template'
    pdf = pdfkit.from_url(projectUrl, False)
    # Generate download
    response = HttpResponse(pdf,content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="ourcodeworld.pdf"'

    return response
'''


class AddCommentView(CreateView):
    comment_form_class = AddCommentForm
    template_name = 'PDFGenerate/base.html'

    def post(self, request):
        comment_form = self.comment_form_class(request.POST, prefix='comment', auto_id=False)

        if comment_form.is_valid():
            comment = Comment()
            comment.message = comment_form.cleaned_data['message']

            return HttpResponseRedirect('/thanks/')
        return render(request, self.template_name, {
            "comment_form": comment_form
        })


def forms(request):
    template_name = 'PDFGenerate/base.html'
    if request.method == 'POST':
        form = AddCommentForm(request.POST, request.FILES)
        id_form=AddIdForm(request.POST,request.FILES)
        #print (id_form)
        if form.is_valid():
            num_results = data.objects.filter(email=form.cleaned_data['message']).count()

            if num_results!=0:
                return render(request, 'PDFGenerate/pdf_form.html', {
                    "field": pdf
                })
            else:
                return HttpResponseRedirect('/form/wrong')

        elif id_form.is_valid():
            post = id_form.save(commit=False)
            post.user = request.user
            post.save()
        else:
            return HttpResponseRedirect('/thanks/')
    else:
        form = AddCommentForm()
    return render(request, template_name, {
        "comment_form": form
    })

'''

class form(CreateView):
    model = pdf
    fields =['name','org','talk','person_image']

    comment_form_class = AddCommentForm
    id_form_class = AddIdForm
    template_name = 'PDFGenerate/base.html'

    def post(self, request):
        comment_form = self.comment_form_class(request.POST, prefix='comment', auto_id=False)
        id_form = self.id_form_class(request.POST, prefix='comment', auto_id=True)
        pf = AddIdForm(request.POST, request.FILES)
        print(pf)

        if pf.is_valid():
            pf.save()
            return HttpResponseRedirect('/thanks/')
            #pf.name=id_form.cleaned_data['name']
            #pf.org= id_form.cleaned_data['org']
            #pf.talk = id_form.cleaned_data['talk']
            #pf.person_image = id_form.cleaned_data[' person_image']
        elif comment_form.is_valid():
            return render(request, 'PDFGenerate/pdf_form.html', {
                "field": pdf
            })

        return render(request, self.template_name, {
            "comment_form": comment_form
        })

'''
class view(generic.ListView):

    template_name = 'PDFGenerate/done.html'
    def get_queryset(self):
        return pdf.objects.all()

class wrong(generic.ListView):

    template_name = 'PDFGenerate/wrong.html'
    def get_queryset(self):
        return pdf.objects.all()


class admin(generic.ListView):
    template_name = 'PDFGenerate/admin.html'
    def get_queryset(self):
        return pdf.objects.all()


def ver(request):

    template_name = 'PDFGenerate/verify.html'

    def choice_func(row):
        q = data.objects.filter(slug=row[0])[0]
        row[0] = q
        return row

    if request.method == 'POST':
        form = AddPasswordForm(request.POST, request.FILES)
        exx=AddExcelForm(request.POST,request.FILES)
        #print (id_form)
        if form.is_valid():

            #print(form.cleaned_data['message'])
            password = form.cleaned_data['password']
            if password == "123":
                return render(request, 'PDFGenerate/admin.html', {
                "object_list": pdf.objects.all()
                })
            else:
                return HttpResponseRedirect('/form/wrong')
        elif exx.is_valid():
            type_list = request.POST.getlist('type', None)

            type=type_list[0]
            data.objects.all().delete()
            filehandle = request.FILES['ex']
            wb = load_workbook(filehandle)
            sheet = wb.get_sheet_by_name('Sheet1')
            print (BASE_DIR)
            for row in sheet.iter_rows('A{}:G{}'.format(sheet.min_row+1, sheet.max_row)):
                d=[]
                for cell in row:
                    d.append(cell.value)

                    #print(cell.value)
                if d[2]==None:
                    company=""
                else:
                    company=d[2]
                if d[3] == None:
                    position = ""
                else:
                    position = d[3]
                if d[4] == None:
                    interest1 = ""
                else:
                    interest1 = d[4]
                if d[5] == None:
                    interest2 = ""
                else:
                    interest2 = d[5]
                if d[6] == None:
                    interest3 = ""
                else:
                    interest3 = d[6]
                datas = data(Number=d[0],FullName=d[1],Company=company,Position=position,Interest1=interest1,Interest2=interest2,Interest3=interest3)
                datas.save()
                #print(d["name"])
            generatePdf(type)

            return render(request, 'PDFGenerate/done.html', {
                "comment_form": form
            })
        else:
            return HttpResponseRedirect('/form/wrong')
    else:
        form = AddPasswordForm
    return render(request, template_name, {
        "comment_form": form
    })



def generatePdf(typ):
    options = {
        'page-height':'4.27in', #actual width
        'page-width':'2.51in', #actual height
        'orientation': 'landscape', #because of landscape mode height and width swaps
        'dpi': 400,
        'margin-top': '0mm',
        'margin-right': '0mm',
        'margin-bottom': '0mm',
        'margin-left': '0mm',
        'encoding': "UTF-8"
    }
    img2 = BASE_DIR + "/PDFGenerate/static/Ted/mini.jpg"
    jqr1 = BASE_DIR + "/PDFGenerate/static/TextResize/jquery-3.2.1.min.js"
    def render_to_pdf(template_src, context_dict,location,img1):
        str1=context_dict.FullName
        size1=len(str1)
        str2=context_dict.Position + context_dict.Company
        size2=len(str2)
        str3=context_dict.Interest1+context_dict.Interest2+context_dict.Interest3
        size3=len(str3)

        if(context_dict.Position=="" and context_dict.Company==""):
            finalStrPos=""
        elif(context_dict.Position=="" and context_dict.Company!=""):
            finalStrPos=context_dict.Company
        elif(context_dict.Position!="" and context_dict.Company==""):
            finalStrPos=context_dict.Position
        else:
            finalStrPos=context_dict.Company+" , "+context_dict.Position
        if(context_dict.Interest1=="" and context_dict.Interest2=="" and context_dict.Interest3==""):
            finalStrInt=""
        else :
            finalStrInt=context_dict.Interest1+" | "+context_dict.Interest2+" | "+context_dict.Interest3

        html = get_template(template_src).render({
            "data": context_dict,
            "img1":img1,
            "img2":img2,
            "jqr1":jqr1,
            "size1": size1,
            "size2": size2,
            "size3": size3,
            "position":finalStrPos,
            "interest":finalStrInt,

        })

        dirs=''.join(PDF_DIRS)

        pdfkit.from_string(html, dirs+location+context_dict.Number + '.pdf', options=options)


    if(typ=="Atendee"):
        location="/Atendee/"
        img1=BASE_DIR+"/PDFGenerate/static/Ted/attendee.png"
        for pdfs in data.objects.all():
            render_to_pdf('PDFGenerate/Ted/Attendee.html', pdfs,location,img1)
    elif(typ=="Crew"):
        location = "/Crew/"
        img1 = BASE_DIR + "/PDFGenerate/static/Ted/crew.png"
        for pdfs in data.objects.all():
            render_to_pdf('PDFGenerate/Ted/Crew.html', pdfs,location,img1)
    elif(typ=="Organizer"):
        location = "/Organizer/"
        img1 = BASE_DIR + "/PDFGenerate/static/Ted/organizer.png"
        for pdfs in data.objects.all():
            render_to_pdf('PDFGenerate/Ted/Organizer.html', pdfs, location,img1)
    elif (typ == "Speaker"):
        location = "/Speaker/"
        img1 = BASE_DIR + "/PDFGenerate/static/Ted/speaker.png"
        for pdfs in data.objects.all():
            render_to_pdf('PDFGenerate/Ted/Speaker.html', pdfs, location,img1)

    elif (typ == "Volunteer"):
        location = "/Volunteer/"
        img1 = BASE_DIR + "/PDFGenerate/static/Ted/organizer.png"
        for pdfs in data.objects.all():
            render_to_pdf('PDFGenerate/Ted/Volunteer.html', pdfs, location,img1)

class gen(generic.ListView):
    template_name = 'PDFGenerate/done.html'


    def render_to_pdf(self,template_src, context_dict,id):
        #template = get_template(template_src)
        #context =Context({"data": context_dict})  # data is the context data that is sent to the html file to render the output.
        #html = template.render(context)  # Renders the template with the context data.
        html=get_template(template_src).render({
            "data": context_dict,
            "dir": STATIC_DIRS
        })
        pdfkit.from_string(html, 'out'+str(id)+'.pdf')
        #pdf = open("out"+id+".pdf")
        #response = HttpResponse(pdf.read(), content_type='application/pdf')  # Generates the response as pdf response.
        #response['Content-Disposition'] = 'attachment; filename=output.pdf'
        #pdf.close()
        # os.remove("out.pdf")  # remove the locally created pdf file.
        #return response  # returns the response.

    def myview(self):
        # Retrieve data or whatever you need
        for pdfs in data.objects.all():
            self.render_to_pdf('PDFGenerate/temp.html', pdfs, pdfs.id)


    def get_queryset(self):
        self.myview()
        return pdf.objects.all()



